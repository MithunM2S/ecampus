from fee import serializers
from rest_framework import viewsets, mixins
from fee import models as fee_model
from rest_framework.views import APIView
from api_authentication.permissions import HasOrganizationAPIKey
from rest_framework.permissions import AllowAny, IsAuthenticated
from django.db.models import Q, F, Sum, Count
from student.models import Profile,ParentDetails
# from student import models as student_models
from rest_framework import response
from master import services
from student import services as student_services
from django_filters.rest_framework import DjangoFilterBackend
from fee import services as fee_services
from rest_framework import status
from datetime import datetime
from openpyxl import Workbook
from django.http import HttpResponse
import os
from base import services as base_services
from master.models import GroupConcat
from django.db.models.functions import Substr

import datetime


class FeeGenericMixinViewSet(mixins.CreateModelMixin,
                                mixins.ListModelMixin,
                                mixins.RetrieveModelMixin,
                                mixins.UpdateModelMixin,
                                viewsets.GenericViewSet):
                                pass


class FeeTypeViewset(viewsets.ModelViewSet):
  queryset = fee_model.FeeType.objects.all()
  serializer_class = serializers.FeeTypeSerializer

  def get_serializer_class(self):
    if self.action == 'create':
      return serializers.FeeTypeCreateSerializer
    if self.action == 'update':
      return serializers.FeeTypeUpdateSerializer
    if self.action == 'retrieve':
      return serializers.FeeTypeDetailSerializer
    if self.action == 'list':
      return serializers.FeeTypeDetailSerializer
    return super(FeeTypeViewset, self).get_serializer_class()

  def perform_create(self, serializer):
    serializer.save(created_by=self.request.user.id)

  def get_queryset(self):
    queryset =  super().get_queryset()
    queryset = queryset.order_by('created_on')
    return queryset
  

class FeeCategoryViewset(viewsets.ModelViewSet):
  queryset = fee_model.FeeCategory.objects.all()
  serializer_class = serializers.FeeCategorySerializer

  def get_serializer_class(self):
    if self.action == 'create':
      return serializers.FeeCategorySerializer
    if self.action == 'update':
      return serializers.FeeCategoryUpdateSerializer
    if self.action == 'retrieve':
      return serializers.FeeCategoryDetailSerializer
    if self.action == 'list':
      return serializers.FeeCategoryDetailSerializer
    return super(FeeCategoryViewset, self).get_serializer_class()

  def perform_create(self, serializer):
    user = self.request.user.id
    serializer.save(created_by=user)
    
  def get_queryset(self):
    queryset =  super().get_queryset()
    queryset = queryset.order_by('created_on')
    print(queryset, '..80')
    return queryset


class FeeToClassViewset(viewsets.ModelViewSet):
  queryset = fee_model.FeeToClass.objects.all()
  serializer_class = serializers.FeeToClassSerializer
  filter_backends = [DjangoFilterBackend]
  filterset_fields = [
              'section',
              'quota',
              'fee_category',
              'fee_type',
              'academic_year',
  ]

  def get_serializer_class(self):
    if self.action == 'create':
      return serializers.FeeToClassCreateSerializer
    if self.action == 'update':
      return serializers.FeeToClassUpdateSerializer
    if self.action == 'retrieve':
      return serializers.FeeToClassDetailSerializer
    if self.action == 'list':
      return serializers.FeeToClassDetailSerializer
    return super(FeeToClassViewset, self).get_serializer_class()

  def get_serializer(self, *args, **kwargs):
      if "data" in kwargs:
        request_data = kwargs["data"]
        if isinstance(request_data, list):
            kwargs["many"] = True

      return super(FeeToClassViewset, self).get_serializer(*args, **kwargs)

  def get_queryset(self):
      filter_by = self.request.query_params.get('type', None)
      queryset = super().get_queryset()
      if filter_by == 'student':
        queryset = queryset.filter(student__isnull=False)
      elif filter_by == 'class':
        # queryset = queryset.filter(class_name__isnull=False)
        queryset = queryset.filter(class_name_id=self.request.query_params['class_name'])

        '''below two lines will filter out the rows
          based on the current month bcz every month
          you having the same fee and client asked to
          show the monthly fee for only one row out of
          12 rows along with the row for anual.'''

        now = datetime.datetime.now().month
        queryset = queryset.filter(Q(month__month=now) | Q(month=None))
      else:
        queryset = queryset
      return queryset

  def perform_create(self, serializer):
    '''
    After we get the validated from the serializer,
    we convert to dict to check if the fee_type is montly
    if the fee type is montly we take start date and end date
    and we create a FeeToClass instances.
    
    '''
    data = dict(serializer.validated_data) #this is a validated data from the serialzier 
    user = self.request.user.id
    if data['fee_type'].fee_type == "monthly":
        start_date = data['start_date'] #date object
        end_date = data['end_date'] #date object
        num_of_months = fee_services.calculate_month_difference(start_date, end_date)
        print(num_of_months)
        if num_of_months == 0: 
          num_of_months = 1 
          '''checking if the start and end date comes on the same month,
          we are collecting fee only for that month therefore making it 1.'''
        
        month = start_date
        if num_of_months > 1:
          for i in range(0, num_of_months):
            
            fee_model.FeeToClass.objects.create(
              start_date = start_date,
              end_date = end_date,
              class_name = data['class_name'],
              section = data['section'],
              quota = data['quota'],
              fee_category = data['fee_category'],
              fee_type = data['fee_type'],
              new_student_amount = data['new_student_amount'],
              old_student_amount = data['old_student_amount'],
              created_by = user,
              academic_year = data['academic_year'],
              month = month   
            )
            if month.month == 12:
              month = month.replace(month = 1, year = month.year + 1)
            else:
              month = month.replace(month = month.month + 1)
        serializer.save(created_by=user, month=month)   
    else:
      serializer.save(created_by=user)
      
  # def perform_update(self, serializer):
  #   return super().perform_update(serializer)
  def perform_update(self, serializer):
    '''
    Monthly fee updation,
    since annual fees updation happens only on one instance,
    montly fees happens over all the instance of the month.
    let's say you have 12 montly fee_to_class instance it 
    updates over all.
    '''
    user = self.request.user.id
    data = serializer.validated_data #contains all the validated data of the form 
    fee_type_instance = data['fee_type']
    if fee_type_instance.fee_type == "monthly":
      class_instance = data['class_name']
      quota = data['quota']
      fee_category = data['fee_category']
      start_date = data['start_date']
      end_date = data['end_date']
      queryset = fee_model.FeeToClass.objects.filter(class_name__id = class_instance.id,   
                                              quota__id = quota.id, 
                                              fee_type__id=fee_type_instance.id,
                                              fee_category__id = fee_category.id).order_by('created_on')
      num_of_months = fee_services.calculate_month_difference(start_date=start_date, end_date=end_date) + 1
      start_date = data['start_date']
      for instance in queryset:
        if num_of_months > 0:
          instance.old_student_amount, instance.new_student_amount = data['old_student_amount'], data['new_student_amount'] 
          instance.month = start_date
          instance.save()
    
        
          if start_date.month == 12:
            start_date = start_date.replace(month = 1, year = start_date.year + 1)
          else:
            start_date = start_date.replace(month = start_date.month + 1)
          num_of_months -= 1
        else:
          instance.delete()
      
      while num_of_months > 0:
        '''
            There may be cases where we have to create more
            instances, let say there were only 5 instances before 
            now if we change start date and end date, and if we have to create 10 instances 
            5 instances will be already updated by the above code, but now we have to create 5 more 
            new instances. This part will do it.
        '''
        
        fee_model.FeeToClass.objects.create(start_date=data['start_date'],
                                              end_date=data['end_date'],
                                              class_name=data['class_name'],
                                              section=data['section'],
                                              quota=data['quota'],
                                              fee_category=data['fee_category'],
                                              fee_type=data['fee_type'],
                                              new_student_amount=data['new_student_amount'],
                                              old_student_amount=data['old_student_amount'],
                                              created_by=user,
                                              academic_year=data['academic_year'],
                                              month=start_date)
         
       
        if start_date.month == 12:
            start_date = start_date.replace(month = 1, year = start_date.year + 1)
         
        else:
            start_date = start_date.replace(month = start_date.month + 1)
         
        num_of_months -=1
        
    else:
      return super().perform_update(serializer)
    
    
class PaymentModeViewset(viewsets.ModelViewSet):
  queryset = fee_model.PaymentMode.objects.all()
  serializer_class = serializers.PaymentModeSerializer

  def get_serializer_class(self):
    if self.action == 'retrieve':
      return serializers.PaymentModeDetailSerializer
    if self.action == 'list':
      return serializers.PaymentModeDetailSerializer
    return super(PaymentModeViewset, self).get_serializer_class()

  def perform_create(self, serializer):
    user = self.request.user.id
    serializer.save(created_by=user)
  
  def get_queryset(self):
    queryset =  super().get_queryset()
    queryset = queryset.order_by('created_on')
    return queryset


# Fee Concession ViewSet
class FeeConcessionViewset(FeeGenericMixinViewSet):
  queryset = fee_model.FeeConcession.objects.all()
  serializer_class = serializers.FeeConcessionSerializer
  http_method_names = ['get', 'post', 'update']

  def get_serializer(self, *args, **kwargs):
      if "data" in kwargs:
        request_data = kwargs["data"]
        if isinstance(request_data, list):
            kwargs["many"] = True

      return super(FeeConcessionViewset, self).get_serializer(*args, **kwargs)

  def get_serializer_class(self):
    if self.action == 'create':
      return serializers.CreateFeeConcessionSerializer
    # if self.action == 'list':
    #   return serializers.PaymentModeDetailSerializer
    return super(FeeConcessionViewset, self).get_serializer_class()

  def perform_create(self, serializer):
    user = self.request.user.id
    serializer.save(created_by=user)


# Fetching Fee Structure of Student
class FetchFees(APIView):
    permission_classes = [HasOrganizationAPIKey, IsAuthenticated]

    def get(self, request, student_id, fee_category_id, *args, **kwargs):
      fees_response = {}
      run_academic_year = services.get_academic_years_key_value('running')[0]
      up_academic_year = services.get_academic_years_key_value('upcoming')[0]
      academicYear = request.GET.get('academicYear', None)
      # fee_category = request.GET.get('feeCategory', None)
      
      if not academicYear:
        academicYear = run_academic_year
      student = student_services.get_student(student_id)
      
      if student:
        studentId = student.id
        if student_services.get_student_state(student_id) ==   'new_student': #get student state helps to identify old and new student
          fee_amount_field = 'new_student_amount'
        else :
          fee_amount_field = 'old_student_amount'
        
        queryset = fee_model.FeeToClass.objects.select_related('class_name', 'section', 'quota', 'fee_category', 'fee_type').values(
          feetoClassId=F('id'),
          academicYear=F('academic_year'),
          className=F('class_name__class_name'),
          classId=F('class_name__id'),
          SectionName=F('section__section_name'),
          SectionId=F('section__id'),
          quotaName=F('quota__name'),
          quotaId=F('quota__id'),
          feeCategoryName=F('fee_category__fee_category'),
          feeCategoryId=F('fee_category__id'),
          feeType=F('fee_type__fee_type'),
          feeTypeName=F('fee_type__fee_type_name'),
          feeTypeId=F('fee_type__id'),
          feeAmount=F(fee_amount_field),
          mont=F('month')
        ).filter(
          (Q(class_name=student.class_name) | Q(student=student.id)),
          quota=student.quota,
          ).order_by('id')
        
        '''
        this if statement will avail to filter conditionally on fee_category
        '''

        if int(fee_category_id):
          queryset = queryset.filter(
          fee_category=fee_category_id).order_by('id')
        
        
        fees_response = queryset 
       
        # print(queryset)
        for key,fees in enumerate(fees_response):
          concessionObject = get_concession(student_id, fees.get('feetoClassId', 0), academicYear)
          balanceObject = get_balance(student_id, fees.get('feetoClassId', 0), academicYear)
          concessionAmount = concessionObject.get('concession_amount') if concessionObject else 0
          fees_response[key]['concessionAmount'] = concessionAmount
          fees_response[key]['concessionId'] = concessionObject.get('id') if concessionObject else None
          fees_response[key]['balanceAmount'] = float(balanceObject.balance_amount) - float(concessionAmount) if balanceObject else float(fees.get('feeAmount')) - float(concessionAmount)
      return response.Response(fees_response)


# Fee Collection ViewSet
class FeeCollectionViewset(FeeGenericMixinViewSet):
  queryset = fee_model.FeeCollection.objects.all()
  serializer_class = serializers.FeeCollectionSerializer
  http_method_names = ['get', 'post']
  filter_backends = [DjangoFilterBackend]
  filterset_fields = [
              'fee_to_class'
  ]

  bill_id = None
  def get_serializer(self, *args, **kwargs):
      if "data" in kwargs:
        request_data = kwargs["data"]
        # check if many is required
        if isinstance(request_data, list):
            kwargs["many"] = True

      return super(FeeCollectionViewset, self).get_serializer(*args, **kwargs)

  def get_serializer_class(self):
    if self.action == 'create' or self.action == 'update':
      return serializers.CreateFeeCollectionSerializer
    return super(FeeCollectionViewset, self).get_serializer_class()

  def create(self, request, *args, **kwargs):
      context = {}
      serializer = self.get_serializer(data=request.data)
      serializer.is_valid(raise_exception=True)
      self.perform_create(serializer)
      context['id'] = self.bill_id
      headers = self.get_success_headers(serializer.data)
      return response.Response(context, status=status.HTTP_201_CREATED)

  def perform_create(self, serializer):
    user = self.request.user.id
    serializer_row = 0
    total_paid_amount = []
    postData = self.request.data
    for data in serializer.data:
      fee_amount = data.get('fee_amount', None)
      paid_amount = data.get('paid_amount', None)
      student_Id = data.get('student_id', None)
      fee_to_class_id = data.get('fee_to_class', None)
      academic_year = data.get('academic_year', None)
      concession_id = data.get('concession', None)
      payment_mode = data.get('payment_mode', None)
      fee_category_id = postData[serializer_row].get('fee_category', None)
      remarks = data.get('remarks', None)
      is_fee_to_class_object = get_fee_to_class(fee_to_class_id)
      is_balance = get_balance(student_Id, fee_to_class_id, academic_year)
      is_concession = get_concession(student_Id, fee_to_class_id, academic_year)
      fee_amount = is_fee_to_class_object.new_student_amount if student_services.get_student_state(student_Id) == 'new_student' else is_fee_to_class_object.old_student_amount
      concession_amount = float(is_concession.get('concession_amount', 0)) if is_concession else 0
      serializer.validated_data[serializer_row]['fee_amount'] = fee_amount
      if is_balance:
        if not is_balance.balance_amount > 0:
          del serializer.validated_data[serializer_row]
          continue
        if float(paid_amount) > float(is_balance.balance_amount):
          final_balance_amount = 0
          serializer.validated_data[serializer_row]['paid_amount'] = float(is_balance.balance_amount) - float(concession_amount)
        else:
          final_balance_amount = float(is_balance.balance_amount) - float(paid_amount) - float(concession_amount)
      else:
        if float(paid_amount) > float(fee_amount):
          final_balance_amount = 0
          serializer.validated_data[serializer_row]['paid_amount'] = float(fee_amount) - float(concession_amount)
        else:
          final_balance_amount = float(fee_amount) - float(paid_amount) - float(concession_amount)
      serializer.validated_data[serializer_row]['balance_amount'] = final_balance_amount
      if is_concession:
          final_concession_amount = concession_amount
          update_concession(is_concession.get('id', 0))
      else:
        final_concession_amount = 0
      serializer.validated_data[serializer_row]['concession_amount'] = final_concession_amount
      total_paid_amount.append(serializer.validated_data[serializer_row]['paid_amount'])
      serializer_row += 1
    collections = serializer.save(created_by=user)
    if collections:
      master_collection = {}
      master_collection['student'] = student_services.get_student(student_Id)
      master_collection['payment_mode'] = fee_services.get_payment_mode(payment_mode)
      master_collection['remarks'] = remarks
      collection_ids = ",".join([str(collection.id) for collection in collections])
      master_collection['fee_collections'] = collection_ids
      master_collection['total_paid_amount'] = sum(total_paid_amount)
      uuid = services.get_timestamp()
      master_collection['bill_number'] = fee_model.FeeMasterCollection.objects.filter(
      fee_category=fee_category_id,
      academic_year=academic_year).count() + 1
      master_collection['uuid'] = uuid
      master_collection['fee_category'] = fee_model.FeeCategory.objects.get(id=fee_category_id)
      master_collection['academic_year'] = academic_year
      master_collection['created_by'] = self.request.user.id
      fee_master = fee_model.FeeMasterCollection.objects.create(**master_collection)
      self.bill_id = fee_master.id


class FeeCollectionReport(mixins.ListModelMixin, viewsets.GenericViewSet):
  queryset = Profile.objects.all()
  serializer_class = serializers.StudentSerializer
  http_method_names = ['get']
  filter_backends = [DjangoFilterBackend]
  filterset_fields = [
              'section',
              'id',
              'class_name',
  ]

  def list(self, request, *args, **kwargs):
    collection_report = generate_collection_report(self.request, self.queryset, self.filter_queryset, self.paginate_queryset)
    response = super().list(request, args, kwargs)
    response.data['results'] = collection_report
    return response


# Search Fee to class by Search
class SearchFeeToClass(APIView):
  permission_classes = [HasOrganizationAPIKey, IsAuthenticated]
  
  def get(self, request, *args, **kwargs):
    student_id = request.GET.get('studentId', None)
    fees_to_class_response = {}
    if student_id:
      queryset = fee_model.FeeToClass.objects.select_related('class_name', 'section', 'quota', 'fee_category', 'fee_type').values(
      feetoClassId=F('id'),
      feeCategoryName=F('fee_category__fee_category'),
      feeCategoryId=F('fee_category__id'),
      feeType=F('fee_type__fee_type'),
      feeTypeName=F('fee_type__fee_type_name'),
      feeNewStudentAmount=F('new_student_amount'),
      feeOldStudentAmount=F('old_student_amount'),
      ).filter(
      Q(class_name__in=Profile.objects.values('class_name').filter(id=student_id, is_active=True)) |
      Q(student=student_id),
      )
      fees_to_class_response=queryset
    return response.Response(fees_to_class_response)

def get_concession(student_id, fee_to_class_id, academicYear):
  concessionObject = fee_model.FeeConcession.objects.values('id', 'concession_amount').filter(
    student_id=student_id,
    fee_to_class=fee_to_class_id,
    academic_year=academicYear,
    is_valid=True
  )
  return concessionObject[0] if concessionObject else 0

def get_balance(studentId, fee_to_class_id, academicYear):
  balanceObject = fee_model.FeeCollection.objects.filter(
  student_id=studentId,
  fee_to_class=fee_to_class_id,
  academic_year=academicYear).order_by('-id')
  return balanceObject[0] if balanceObject else 0

def get_student_concession(studentId, fee_to_class_id, academicYear, fee_type):
  studentConcessionObject = fee_model.FeeCollection.objects.select_related('fee_to_class').values('concession_amount').filter(
  student_id=studentId,
  fee_to_class=fee_to_class_id,
  academic_year=academicYear,
  fee_to_class__fee_type=fee_type).annotate(total_concession_amount=Sum('concession_amount'))
  return studentConcessionObject[0] if studentConcessionObject else 0

def update_concession(cid):
  fee_model.FeeConcession.objects.filter(id=cid).update(is_valid=False)

def get_fee_to_class(fid):
  try:
    return fee_model.FeeToClass.objects.get(id=fid)
  except Exception as e:
    return None

def get_total_paid(studentId, fee_to_class_id, academicYear):
  paidObject = fee_model.FeeCollection.objects.all().filter(student_id=studentId,
  fee_to_class_id =fee_to_class_id,
  academic_year=academicYear).aggregate(total_paid = Sum('paid_amount'))

#  academic_year=academicYear).annotate(total_paid=Sum('paid_amount'))
  # print(fee_to_class_id) 

  if paidObject['total_paid'] ==  None:
    paidObject['total_paid']=0
  # print(paidObject)
  abc=paidObject['total_paid']
  return abc if paidObject else 0


  # paidObject =0
  

class FeeMasterCollectionViewset(mixins.ListModelMixin, mixins.RetrieveModelMixin, viewsets.GenericViewSet):
  lookup_fields = ('uuid')
  queryset = fee_model.FeeMasterCollection.objects.all()
  serializer_class = serializers.FeeMasterCollectionSerializer
  filter_backends = [DjangoFilterBackend]
  filterset_fields = {
              'student':['exact'],
              'payment_mode':['exact'],
              'fee_category':['exact'],
              'bill_number':['exact'],
              'created_on':['gte', 'lte'],
  }
  '''this function will return the response with adding grand total to it.
  '''
  def list(self, request, *args, **kwargs):
    response = super().list(request, args, kwargs)
    grand_total = 0    
    for person in response.data['results']:
      grand_total+=float(person['total_paid_amount'])
    response.data['results']=[{'collectionData':response.data['results']},{'final_total_paid_amount':grand_total}]
    return response

class DailyReportViewset(mixins.ListModelMixin, viewsets.GenericViewSet):
  lookup_fields = ('uuid')
  queryset = fee_model.FeeMasterCollection.objects.all()
  serializer_class = serializers.DailyReportSerializer
  filter_backends = [DjangoFilterBackend]
  filterset_fields = {
              'fee_category': ['exact'],
              'created_on': ['gte', 'lte'],
  }


class DailyReportXlsx(APIView):
  permission_classes = [HasOrganizationAPIKey, IsAuthenticated]

  def get(self, request, *args, **kwargs):

    fee_category = self.request.query_params.get('fee_category', None)
    created_on__gte = self.request.query_params.get('created_on__gte', None)
    created_on__lte = self.request.query_params.get('created_on__lte', None)
    if fee_category and not created_on__lte and not created_on__gte:
      queryset = fee_model.FeeMasterCollection.objects.filter(Q(fee_category=fee_category))
    elif not fee_category and created_on__lte and created_on__gte:
        queryset = fee_model.FeeMasterCollection.objects.filter(Q(created_on__gte=created_on__gte),
                                                                Q(created_on__lte=created_on__lte))
    elif fee_category and created_on__lte and created_on__gte:
      queryset = fee_model.FeeMasterCollection.objects.filter(Q(fee_category=fee_category),
                                                              Q(created_on__gte=created_on__gte),
                                                              Q(created_on__lte=created_on__lte))

    else:
      queryset = fee_model.FeeMasterCollection.objects.filter()

    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Daily Report'

    # Define the titles for columns
    columns = [
      'Date',
      'Bill Number',
      'Student ID',
      'Student Name',
      'Class',
      'Total Paid Amount',
      'Payment Mode',
      'Fee Category',
      'Company Name'
    ]
    max_count = 0
    for fee in queryset:
      fee_collections = fee_model.FeeCollection.objects.select_related('fee_to_class',
                                                                       'fee_to_class__fee_type').values(
        feeName=F('fee_to_class__fee_type__fee_type_name'),
        paidAmount=F('paid_amount'),

      ).filter(
        id__in=[cid for cid in fee.fee_collections.split(",")]
      ).count()
      max_count = max(max_count, fee_collections)

    for i in range(max_count):
        columns.append('Fee Type')
        columns.append('Amount Paid')
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
      cell = worksheet.cell(row=row_num, column=col_num)
      cell.value = column_title
    total_amount = sum(queryset.values_list('total_paid_amount', flat=True))
    # Iterate through all movies
    for fee_obj in queryset:
      row_num += 1
      fee_collections = fee_model.FeeCollection.objects.select_related('fee_to_class',
                                                                                'fee_to_class__fee_type').values(
        feeName=F('fee_to_class__fee_type__fee_type_name'),
        paidAmount=F('paid_amount'),

      ).filter(
        id__in=[cid for cid in fee_obj.fee_collections.split(",")]
      )
      payment_mode = services.get_related(fee_obj.payment_mode, 'name')[0]['name']
      fee_category = services.get_related(fee_obj.fee_category, 'fee_category')
      companyName = services.get_related(fee_obj.fee_category, 'company_name')
      if fee_category:
        fee_category = fee_category[0]['name']
        # companyName = fee_category[0]['name']
      else:
        fee_category = ""
        # companyName = ""
      # Define the data for each cell in the row
      
      if companyName:
        companyName = companyName[0]['name']
        # companyName = fee_category[0]['name']
      else:
        companyName = ""
        # companyName = ""
      # Define the data for each cell in the row

      row = [
        fee_obj.created_on,
        fee_obj.bill_number,
        fee_obj.student.student_id,
        fee_obj.student.first_name,
        fee_obj.student.class_name.class_name,
        fee_obj.total_paid_amount,
        payment_mode,
        fee_category,
        companyName
      ]
      for fee_collection in fee_collections:
        row.append(fee_collection['feeName'])
        row.append(fee_collection['paidAmount'])
      # Assign the data for each cell of the row
      for col_num, cell_value in enumerate(row, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = cell_value
    cell = worksheet.cell(row=row_num + 1, column=1)
    cell.value = 'Daywise Total'
    cell = worksheet.cell(row=row_num+1, column=2)
    cell.value = total_amount
    dir_path = os.path.join("media/", "files/fee/")
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)
    file_path = os.path.join(dir_path, "daily-fee-report.xlsx")
    workbook.save(file_path)

    return response.Response({"path": base_services.get_hostname(request) + "/" + file_path})


def generate_collection_report(request, student_queryset, filter_queryset=None, paginate_queryset=None):
  filter_fee_category = request.query_params.get('fee_category', None)
  filter_academic_year = request.query_params.get('academic_year', None)
  if not filter_academic_year:
    filter_academic_year = services.get_academic_years_key_value('running')[0]
  fee_category_obj = fee_model.FeeCategory.objects
  if filter_fee_category:
    fee_category_obj = fee_category_obj.filter(id=filter_fee_category)
  else:
    fee_category_obj = fee_category_obj.all()
  fee_category_collection_report = {}
  fee_category_collection_report['collectionData'] = []
  final_total_fee_amount, final_total_paid_amount, final_total_concession_amount,final_total_balance_amount = 0, 0, 0, 0
  for feeCatKey, feeCat in enumerate(fee_category_obj, 0):
    report_data = {}
    report_data['feeCategoryId'] = feeCat.id
    report_data['feeCategoryName'] = feeCat.fee_category
    report_data['feeCategoryCompanyName'] = feeCat.company_name
    # report_data['feeCategoryphoto'] = feeCat.logo
    if filter_queryset and paginate_queryset:
      students = paginate_queryset(filter_queryset(student_queryset))
    else:
      students = students
    fee_category_collection = []
    for student in students:
      if student_services.get_student_state(student.id) == 'new_student':
        fee_amount_field = 'new_student_amount'
      else :
        fee_amount_field = 'old_student_amount'
      fee_types = fee_model.FeeToClass.objects.select_related('student_id','fee_type').values(
          feeTypeId=F('fee_type__id'),
          feeToClassId=F('id'),
          feeName=F('fee_type__fee_type_name'),
          feeAmount=F(fee_amount_field)).filter(
            fee_category=feeCat.id,
            class_name=student.class_name,
            academic_year=filter_academic_year
          ).order_by('fee_type__id')
      fee_type_list = []
      student_fee_typewise_total_list = []
      student_fee_typewise_total = {}
      student_total_fee_amount, student_total_paid_amount, student_total_concession_amount,student_total_balance_amount = 0, 0, 0, 0
      for row, fee_type in enumerate(fee_types, 0):
        # print(fee_type.get('feeToClassId'))
        paid_amount = get_total_paid(student.id, fee_type.get('feeToClassId'), filter_academic_year)
        # print(paid_amount)
        concession_amount = get_student_concession(student.id, fee_type.get('feeToClassId'), filter_academic_year, fee_type.get('feeTypeId'))
        balance_amount = get_balance(student.id, fee_type.get('feeToClassId'), filter_academic_year)
        concession_amount = int(concession_amount.get('total_concession_amount', 0)) if concession_amount else 0
        balance_amount = int(balance_amount.balance_amount) if balance_amount else int(fee_type.get('feeAmount'))
        fee_types[row]['feeAmount'] = int(fee_type.get('feeAmount'))
        fee_types[row]['paidAmount'] = int(paid_amount)
        fee_types[row]['concessionAmount'] = concession_amount
        fee_types[row]['balanceAmount'] = balance_amount
        report_data['totalFeeAmount'] = report_data.get('totalFeeAmount', 0) + int(fee_type.get('feeAmount'))
        report_data['totalPaidAmount'] = report_data.get('totalPaidAmount', 0) + int(paid_amount)
        # print(fee_types)
        report_data['totalConcessionAmount'] = report_data.get('totalConcessionAmount', 0) + concession_amount
        report_data['totalBalanceAmount'] = report_data.get('totalBalanceAmount', 0) + balance_amount
        fee_type_list.append(fee_types[row])
        # print(paid_amount)
        student_fee_typewise_total['totalFeeAmount'] = student_fee_typewise_total.get('totalFeeAmount', 0) + fee_types[row]['feeAmount']
        student_fee_typewise_total['totalPaidAmount'] =  student_fee_typewise_total.get('totalPaidAmount', 0) + fee_types[row]['paidAmount']
        student_fee_typewise_total['totalConcessionAmount'] = student_fee_typewise_total.get('totalConcessionAmount', 0) + fee_types[row]['concessionAmount']
        student_fee_typewise_total['totalBalanceAmount'] = student_fee_typewise_total.get('totalBalanceAmount', 0) + fee_types[row]['balanceAmount']
      student_fee_typewise_total_list.append(student_fee_typewise_total)
      fee_category_collection.append({
        'id':student.id,
        'studentId':student.student_id,
        'studentName':student.first_name,
        'className': services.get_related(student.class_name, 'class_name'),
        'section': services.get_related(student.section, 'section_name'),
        'feeType': fee_type_list,
        'collectionTotal': student_fee_typewise_total_list
      })
    report_data['feeCategoryCollection'] = fee_category_collection
    final_total_fee_amount = final_total_fee_amount + report_data.get('totalFeeAmount', 0)
    final_total_paid_amount = final_total_paid_amount + report_data.get('totalPaidAmount', 0)
    final_total_concession_amount = final_total_concession_amount + report_data.get('totalConcessionAmount', 0)
    final_total_balance_amount = final_total_balance_amount + report_data.get('totalBalanceAmount', 0)
    fee_category_collection_report['collectionData'].append(report_data)
    # fee_category_collection_report.append(report_data)
  final_total_data = {
    'finalTotalFeeAmount':final_total_fee_amount,
    'finalTotalPaidAmount': final_total_paid_amount,
    'finalTotalConcessionAmount': final_total_concession_amount,
    'finalTotalBalanceAmount': final_total_balance_amount
  }
  fee_category_collection_report['FinalTotalData'] = final_total_data
  return [fee_category_collection_report]


# Fee Dashboard
class FeeDashboard(APIView):
    permission_classes = [HasOrganizationAPIKey, IsAuthenticated]

    def get(self, request, *args, **kwargs):
      return response.Response(generate_dashboard_report(request)) 


# Generation fee dashboard report
def generate_dashboard_report(request):
  filter_class_group = request.query_params.get('classGroup', None)
  filter_class_name = request.query_params.get('className', None)
  filter_section = request.query_params.get('section', None)
  filter_fee_category = request.query_params.get('feeCategory', None)
  filter_fee_type = request.query_params.get('feeTypes', None)
  fee_dashboard_report = {}
  fee_dashboard_report['dashboardData'] = []
  grand_total = 0
  fee_category_obj = fee_model.FeeCategory.objects.all()
  student_profile_queryset = Profile.objects.values('id')
  if filter_class_group:
    student_profile_queryset = student_profile_queryset.filter(class_group=filter_class_group)
  if filter_class_name:
    student_profile_queryset = student_profile_queryset.filter(class_name=filter_class_name)
  if filter_section:
    student_profile_queryset = student_profile_queryset.filter(section=filter_section)
  if filter_fee_category:
    fee_category_obj = fee_category_obj.filter(id=filter_fee_category)
  fee_final_total_dict = {}
  for fee_category in fee_category_obj:
    fee_dashboard = {}
    fee_dashboard['feeCategoryCompanyName'] = fee_category.company_name
    fee_dashboard['feeCategory'] = fee_category.fee_category
    fee_dashboard['feeCategoryId'] = fee_category.id
    fee_dashboard['feeCategoryTotal'] = 0
    fee_types_list= []
    fee_type_queryset = fee_model.FeeType.objects.values(
      typeName=F('fee_type_name'),
      typeId=F('id')).order_by('id')
    if filter_fee_type:
      fee_type_queryset = fee_type_queryset.filter(id=filter_fee_type)
    fee_total_dict = {}
    for fee_type_ in fee_type_queryset:
      fee_to_class_queryset = fee_model.FeeToClass.objects.values('id').filter(
        fee_category=fee_category.id,
        fee_type=fee_type_.get('typeId', None))
      if not fee_to_class_queryset.exists():
        continue
      fee_to_class_ids = [a.get('id') for a in fee_to_class_queryset]
      payment_mode_list = []
      payment_mode_total = 0
      for ptype in fee_model.PaymentMode.objects.all().order_by('id'):
        fee_collection_queryset = fee_model.FeeCollection.objects.filter(
          payment_mode=ptype.id,
          fee_to_class__in=fee_to_class_ids,
          student_id__in=student_profile_queryset).aggregate(totalSum=Sum('paid_amount'))
        to = fee_collection_queryset.get('totalSum', None) if fee_collection_queryset.get('totalSum', None) else 0
        payment_mode_total = payment_mode_total + to
        if fee_total_dict.get(ptype.name, None): 
          fee_total_dict[ptype.name] = fee_total_dict[ptype.name] + to
        else:
          fee_total_dict[ptype.name] = to
        if fee_final_total_dict.get(ptype.name, None):
          fee_final_total_dict[ptype.name] = fee_final_total_dict[ptype.name] + to
        else:
          fee_final_total_dict[ptype.name] = to
        grand_total = grand_total + to
        fee_dashboard['feeCategoryTotal'] = fee_dashboard['feeCategoryTotal'] + to
        payment_mode_list.append({ptype.id: {"name":ptype.name, "total": to }})
      fee_type_['paymentModes'] = payment_mode_list
      fee_type_['totalOfPaymentMode'] = payment_mode_total
      fee_types_list.append(fee_type_)
    fee_dashboard['feeTypes'] = fee_types_list
    fee_dashboard['paymentModeWiseTotal'] = fee_total_dict
    fee_dashboard_report['dashboardData'].append(fee_dashboard)
  if fee_dashboard_report:
    fee_final_total_dict['grandTotal'] = grand_total
    fee_dashboard_report['finalTotalData'] = fee_final_total_dict
    fee_dashboard_report['paymentModes'] = fee_model.PaymentMode.objects.values('id', 'name').order_by('id')
  return fee_dashboard_report


# Bill to Bill Report
class BillToBillReport(APIView):
    permission_classes = [HasOrganizationAPIKey, IsAuthenticated]

    def get(self, request, *args, **kwargs):
      print('called .. ?')
      query_fee_category = request.query_params.get('feeCategory', None)
      query_from_date = request.query_params.get('fromDate', None)
      query_to_date = request.query_params.get('toDate', None)
      bill_to_bill_report = {}
      if request.user.has_perm('fee.view_feemastercollection') and query_from_date and query_to_date:
        fee_master_collecion = fee_model.FeeMasterCollection.objects.values(
          date=F('created_on')
        ).filter(
          created_on__gte=query_from_date,
          created_on__lte=query_to_date,
          fee_category=query_fee_category
        ).annotate(
          billNumber=GroupConcat('bill_number'),
          billCount=Count('created_on'),
          totalAmountCollected=Sum('total_paid_amount')
        )
        billTotalAmount = 0
        for row,value in enumerate(fee_master_collecion):
          billNumbers = value.get('billNumber').split(',')
          fee_master_collecion[row]['fromBillNumber'] = int(billNumbers[0])
          fee_master_collecion[row]['toBillNumber'] = int(billNumbers[-1])
          billTotalAmount = billTotalAmount + value.get('totalAmountCollected')
          del fee_master_collecion[row]['billNumber']
        bill_to_bill_report['billData'] = fee_master_collecion
        if billTotalAmount:
          bill_to_bill_report['grandTotalCollectedAmount'] = billTotalAmount
      return response.Response(bill_to_bill_report)
